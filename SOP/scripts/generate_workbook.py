#!/usr/bin/env python3
"""Generate an operational cost-estimate workbook from normalized JSON output."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
SECTION_FONT = Font(bold=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a cost-estimate workbook from normalized JSON."
    )
    parser.add_argument("--input", required=True, help="Path to normalized JSON input")
    parser.add_argument("--output", required=True, help="Path to output xlsx file")
    return parser.parse_args()


def load_payload(path: Path) -> dict:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        width = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, len(value))
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(width + 2, 12), 40
        )


def write_header(worksheet, row_number: int, columns: Iterable[str]) -> None:
    for column_number, name in enumerate(columns, start=1):
        cell = worksheet.cell(row=row_number, column=column_number, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def create_summary_sheet(workbook: Workbook, payload: dict) -> None:
    worksheet = workbook.active
    worksheet.title = "Summary"

    worksheet["A1"] = "Nok Cost Estimate Summary"
    worksheet["A1"].font = Font(size=14, bold=True)

    summary_rows = [
        ("Estimate Name", payload["estimateName"]),
        ("Category", payload["category"]),
        ("Product Name", payload["productName"]),
        ("Region", payload["region"]),
        ("Generated At", payload["generatedAt"]),
        ("Estimate Status", payload["estimateStatus"]),
        ("Total Unit Cost", payload["totalUnitCost"]),
        ("Process Path Summary", payload["processPathSummary"]),
    ]

    worksheet["A3"] = "Metadata"
    worksheet["A3"].fill = SECTION_FILL
    worksheet["A3"].font = SECTION_FONT
    for row_number, (label, value) in enumerate(summary_rows, start=4):
        worksheet.cell(row=row_number, column=1, value=label)
        worksheet.cell(row=row_number, column=2, value=value)

    start_row = len(summary_rows) + 6
    worksheet.cell(row=start_row, column=1, value="Top Cost Drivers")
    worksheet.cell(row=start_row, column=1).fill = SECTION_FILL
    worksheet.cell(row=start_row, column=1).font = SECTION_FONT
    for index, value in enumerate(payload.get("topCostDrivers", []), start=start_row + 1):
        worksheet.cell(row=index, column=1, value=f"Driver {index - start_row}")
        worksheet.cell(row=index, column=2, value=value)

    start_row = start_row + len(payload.get("topCostDrivers", [])) + 2
    worksheet.cell(row=start_row, column=1, value="Open Questions")
    worksheet.cell(row=start_row, column=1).fill = SECTION_FILL
    worksheet.cell(row=start_row, column=1).font = SECTION_FONT
    for index, value in enumerate(payload.get("openQuestions", []), start=start_row + 1):
        worksheet.cell(row=index, column=1, value=f"Question {index - start_row}")
        worksheet.cell(row=index, column=2, value=value)

    autosize_columns(worksheet)


def create_table_sheet(workbook: Workbook, name: str, columns: list[str], rows: list[list]) -> None:
    worksheet = workbook.create_sheet(name)
    write_header(worksheet, 1, columns)

    for row_index, row_data in enumerate(rows, start=2):
        for column_index, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)


def build_step_rows(payload: dict) -> list[list]:
    rows = []
    for step in payload["stepEstimates"]:
        rows.append(
            [
                step["stepName"],
                step["conditionPath"],
                step["included"],
                step["laborMinutes"],
                step["laborRatePerHour"],
                step["laborCostPerUnit"],
                step["partsCostPerUnit"],
                step["handlingCostPerUnit"],
                step["storageCostPerUnit"],
                step["falloutAdjustmentPerUnit"],
                step["totalStepCostPerUnit"],
                step["confidence"],
                step["assumptionSource"],
                step["notes"],
            ]
        )
    return rows


def build_sensitivity_rows(payload: dict) -> list[list]:
    rows = []
    for name, detail in payload.get("sensitivityView", {}).items():
        rows.append([name, detail])
    return rows


def build_memory_rows(payload: dict) -> list[list]:
    rows = []
    for key, items in payload.get("memorySignals", {}).items():
        if isinstance(items, list):
            for item in items:
                rows.append([key, item])
        else:
            rows.append([key, items])
    return rows


def validate_payload(payload: dict) -> None:
    required_top_level = {
        "estimateName",
        "category",
        "productName",
        "region",
        "generatedAt",
        "estimateStatus",
        "totalUnitCost",
        "processPathSummary",
        "topCostDrivers",
        "openQuestions",
        "stepEstimates",
    }
    missing = required_top_level - set(payload)
    if missing:
        raise ValueError(f"Missing required top-level keys: {sorted(missing)}")


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output)

    payload = load_payload(input_path)
    validate_payload(payload)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    create_summary_sheet(workbook, payload)
    create_table_sheet(
        workbook,
        "Step Costs",
        [
            "stepName",
            "conditionPath",
            "included",
            "laborMinutes",
            "laborRatePerHour",
            "laborCostPerUnit",
            "partsCostPerUnit",
            "handlingCostPerUnit",
            "storageCostPerUnit",
            "falloutAdjustmentPerUnit",
            "totalStepCostPerUnit",
            "confidence",
            "assumptionSource",
            "notes",
        ],
        build_step_rows(payload),
    )
    create_table_sheet(
        workbook,
        "Sensitivity",
        ["sensitivityName", "detail"],
        build_sensitivity_rows(payload),
    )
    create_table_sheet(
        workbook,
        "Memory Signals",
        ["signalType", "detail"],
        build_memory_rows(payload),
    )

    workbook.save(output_path)
    print(f"Generated workbook: {output_path}")


if __name__ == "__main__":
    main()
